import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook, load_workbook

def open_csv(file_path):
    try:
        df = pd.read_csv(file_path)
        print(f"File: {file_path}")
        print(df.head())
        return df
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except pd.errors.EmptyDataError:
        print("Error: The file is empty.")
    except pd.errors.ParserError:
        print("Error: The file is not a valid CSV.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def add_values_to_csv(file_path, new_data):
    try:
        df = pd.read_csv(file_path)
        df1 = pd.DataFrame(new_data)
        
        # Ensure new DataFrame df1 has the same columns as the existing DataFrame df
        for col in df.columns:
            if col not in df1.columns:
                df1[col] = None
        
        # Append the new data to the existing DataFrame
        updated_df = pd.concat([df1, df], ignore_index=True)
        
        # Save the updated DataFrame back to the CSV file
        updated_df.to_csv(file_path, index=False)
        
        # Print the first few rows of the updated DataFrame to confirm the addition
        print("New values added to the CSV file:")
        print(updated_df.head())
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except pd.errors.EmptyDataError:
        print("Error: The file is empty.")
    except pd.errors.ParserError:
        print("Error: The file is not a valid CSV.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def add_E_RHE_column(file_path):
    try:
        df = pd.read_csv(file_path)
        required_columns = ['Potential', 'NHE', 'Constant', 'pH', 'Current', 'R']
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Error: The required column '{col}' does not exist in the CSV file.")
        
        # Fill NaNs with the first available value for NHE, pH, Constant, R, and Area 
        df['NHE'] = df['NHE'].fillna(method='ffill')
        df['pH'] = df['pH'].fillna(method='ffill')
        df['Constant'] = df['Constant'].fillna(method='ffill')
        df['R'] = df['R'].fillna(method='ffill')
        df['Area (cm^2)'] = df['Area (cm^2)'].fillna(method='ffill')
        
        # Calculate E(RHE)
        df['E(RHE)'] = df['Potential'] + df['NHE'] + df['Constant'] * df['pH'] - (df['Current'] * df['R'])

        df = df.loc[:, ~df.columns.str.contains('^unnamed')]

         # Rearrange columns to move E(RHE) after Current
        columns = df.columns.tolist()
        columns.remove('E(RHE)')
        columns.insert(columns.index('Current') + 1, 'E(RHE)')
        df = df[columns]
        
        # Save the DataFrame back to the CSV file
        df.to_csv(file_path, index=False)
        
        print("E(RHE) column added to the CSV file:")
        print(df.head())
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except pd.errors.EmptyDataError:
        print("Error: The file is empty.")
    except pd.errors.ParserError:
        print("Error: The file is not a valid CSV.")
    except ValueError as ve:
        print(ve)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def add_current_density_column(file_path):
    try:
        df = pd.read_csv(file_path)
        required_columns = ['Current', 'Area (cm^2)']
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Error: The required column '{col}' does not exist in the CSV file.")
        
        # Calculate Current Density
        df['Current Density'] = df['Current'] / df['Area (cm^2)']

        # Remove unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        # Rearrange columns to move 'Current Density' after 'E(RHE)'
        columns = df.columns.tolist()
        columns.remove('Current Density')
        columns.insert(columns.index('E(RHE)') + 1, 'Current Density')
        df = df[columns]

        df.to_csv(file_path, index=False)
        print('Current Density column added to the CSV file:')
        print(df.head())
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except pd.errors.EmptyDataError:
        print("Error: The file is empty.")
    except pd.errors.ParserError:
        print("Error: The file is not a valid CSV.")
    except ValueError as ve:
        print(ve)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def inspect_columns(file_path):
    try:
        df = pd.read_csv(file_path)
        print(f"Columns in the CSV file '{file_path}':")
        print(df.columns)
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except pd.errors.EmptyDataError:
        print("Error: The file is empty.")
    except pd.errors.ParserError:
        print("Error: The file is not a valid CSV.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def process_all_csv_in_directory(directory_path, new_data):
    # Create a new Excel workbook
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    # Get the list of CSV files and sort them numerically
    csv_files = sorted(
        [filename for filename in os.listdir(directory_path) if filename.endswith(".csv")],
        key=lambda x: int(''.join(filter(str.isdigit, x))) if any(char.isdigit() for char in x) else float('inf')
    )

    # Data to hold for combined plotting
    combined_data = []

    for filename in csv_files:
        file_path = os.path.join(directory_path, filename)
        
        print(f"\nProcessing file: {file_path}")
        
        # Inspect columns to check for correct column names
        inspect_columns(file_path)
        
        # Add new values to the CSV file
        add_values_to_csv(file_path, new_data)
        
        # Add the E(RHE) column to the CSV file
        add_E_RHE_column(file_path)
        
        # Add Current Density column to the CSV file
        add_current_density_column(file_path)
        
        # Extract the required columns
        df = pd.read_csv(file_path)
        required_columns = ['E(RHE)', 'Current Density', 'Scan Rate']
        if all(col in df.columns for col in required_columns):
            # Create a new sheet in the workbook for this CSV file
            sheet_name = os.path.splitext(filename)[0]
            sheet = workbook.create_sheet(title=sheet_name)
            
            # Write the column headers
            for col_num, column_title in enumerate(required_columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            
            # Write the data rows
            for row_num, row in enumerate(df[required_columns].itertuples(index=False, name=None), 2):
                for col_num, cell_value in enumerate(row, 1):
                    sheet.cell(row=row_num, column=col_num, value=cell_value)
            
            # Append the data to combined_data for plotting
            combined_data.append((sheet_name, df[required_columns]))
        
    # Save the workbook to a file
    output_file_path = os.path.join(directory_path, "combined_data.xlsx")
    workbook.save(output_file_path)
    print(f"\nAll data has been extracted and saved to '{output_file_path}'")

    # Now process the saved Excel file to calculate additional metrics and plot data
    process_excel_file(output_file_path, new_data)

def process_excel_file(file_path, new_data):
    # Load the Excel file
    xls = pd.ExcelFile(file_path)

    # Initialize empty lists to store the extracted data
    oxidation_data = []
    reduction_data = []
    sheet_names = []

    # Loop through each sheet in the Excel file
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Check if 'Current Density' column exists in the sheet
        if 'Current Density' in df.columns:
            # Extract data from lines 104 and 303
            data_oxidation = df['Current Density'].iloc[102]
            data_reduction = df['Current Density'].iloc[301]
            
            # Append data to the lists
            oxidation_data.append(data_oxidation)
            reduction_data.append(data_reduction)
            sheet_names.append(sheet_name)

    # Create a new DataFrame for the ECSA worksheet
    ecsa_df = pd.DataFrame({
        'Sheet Name': sheet_names,
        'Current Density Oxidation (A/cm^2)': oxidation_data,
        'Current Density Reduction (A/cm^2)': reduction_data
    })

    # Calculate diffrence between anodic and cathodic for each row and create a new column
    ecsa_df['Current Density (A/cm^2)'] =(ecsa_df['Current Density Oxidation (A/cm^2)'] - ecsa_df['Current Density Reduction (A/cm^2)'])

    # Extract scan rates from sheet names for plotting
    def extract_scan_rate(sheet_name):
        try:
            return float(sheet_name.replace('mV', '').strip())
        except ValueError:
            return np.nan

    ecsa_df['Scan Rate'] = ecsa_df['Sheet Name'].apply(extract_scan_rate)

    # Divide the Scan Rate values by 1000
    ecsa_df['Scan Rate'] = ecsa_df['Scan Rate'] / 1000

    # Reorder the columns
    desired_order = ['Sheet Name', 'Scan Rate', 'Current Density Oxidation (A/cm^2)', 'Current Density Reduction (A/cm^2)', 'Current Density (A/cm^2)']
    ecsa_df = ecsa_df[desired_order]

    # Fitting linear regression
    X = ecsa_df['Scan Rate'].values.reshape(-1, 1)  # Reshape for sklearn
    y = ecsa_df['Current Density (A/cm^2)'].values
    model = LinearRegression()
    model.fit(X, y)

    # Add a new column with the slope value
    ecsa_df['cdl (A cm^-2 V s^-1)'] = model.coef_[0]

    # Add Cs ,R, and Area values to the ECSA DataFrame
    ecsa_df['Cs (mF/cm^2)'] = new_data['Cs(mF/cm^2)'][0]
    ecsa_df['R']=new_data['R'][0]
    ecsa_df['Area (cm^2)']=new_data['Area (cm^2)'][0]

    # Calculate and add the new column 'cdl (mA / Vs^-1)'
    ecsa_df['cdl (mA / Vs^-1)'] = ecsa_df['cdl (A cm^-2 V s^-1)'] * ecsa_df['Area (cm^2)'] * 1000
    # Add cdl(mF) column
    ecsa_df['cdl (mF)']=ecsa_df['cdl (mA / Vs^-1)']

    #calculate ECSA 
    ecsa_df['ECSA (cm^2)']=ecsa_df['cdl (mF)'] / ecsa_df['Cs (mF/cm^2)']
    print('ECSA =',  ecsa_df['ECSA (cm^2)'].head(1))

    # Create a new Excel writer object and add or replace the ECSA sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        ecsa_df.to_excel(writer, sheet_name='ECSA', index=False)

    print("Data has been successfully written to the 'ECSA' worksheet.")

    # Plotting the data
    plt.figure(figsize=(10, 6))
    plt.scatter(ecsa_df['Scan Rate'], ecsa_df['Current Density (A/cm^2)'], color='blue', label='Data points')

    # Plotting the regression line
    y_pred = model.predict(X)
    plt.plot(ecsa_df['Scan Rate'], y_pred, color='red', label=f'Linear fit: y = {model.coef_[0]:.4e}x + {model.intercept_:.4e}')

    # Adding labels and legend
    plt.xlabel('Scan Rate (V/s)')
    plt.ylabel('Current Density (A/cm^2)')
    plt.title('Scan Rate vs Current Density')
    plt.legend()
    plt.grid(True)
    plt.show()

    # Print the slope equation in scientific notation
    print(f"Linear regression equation: y = {model.coef_[0]:.4e}x + {model.intercept_:.4e}")

# Example usage
directory_path = r""#) Enter your path To your folder containing your data 
path_length = r"\combined_data.xlsx" # re-enter the the path but again. this will make the excel sheet 
new_data = {
    'NHE': [0.168],
    'pH': [13],
    'Constant': [0.059],
    'Area (cm^2)': [3.85E-02],
    'R': [63],
    'Cs(mF/cm^2)':[0.040]
}

process_all_csv_in_directory(directory_path, new_data)
